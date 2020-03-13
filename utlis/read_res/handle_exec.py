# -*- coding: UTF-8 -*-
import os
import time
import openpyxl
from openpyxl import workbook
from utlis.file.FileUtil import file_utils
from utlis.Log.logger_config import logger
import ATC


class HandleExec(object):












    def load_excel(self,file_path):
        path  = file_utils.location_file(file_path)
        if not os.path.exists(path):
            logger.info("文件路径不存在:{0}".format(path))
        return openpyxl.load_workbook(path)





    def get_sheet_data(self,index=None):
        '''
        加载所有sheet的内容
        '''
        sheet_name = self.load_excel().sheetnames
        if index == None:
            index = 0
        data = self.load_excel()[sheet_name[index]]
        return data

    def get_cell_value(self, row, cols):
        '''
        获取某一个单元格内容
        '''
        data = self.get_sheet_data().cell(row=row, column=cols).value
        return data

    def get_rows(self):
        '''
        获取行数
        '''
        row = self.get_sheet_data().max_row
        return row

    def get_rows_value(self, row):
        '''
        获取某一行的内容
        '''
        row_list = []
        for i in self.get_sheet_data()[row]:
            row_list.append(i.value)
        return row_list

    def excel_write_data(self,file_paht, row, cols, value):
        '''
        写入数据
        '''
        wb = self.load_excel()
        wr = wb.active
        wr.cell(row, cols, value)
        wb.save(file_utils.location_file(file_paht))

    def get_columns_value(self, key=None):
        '''
        获取某一列得数据
        '''
        columns_list = []
        if key == None:
            key = 'A'
        columns_list_data = self.get_sheet_data()[key]
        for i in columns_list_data:
            columns_list.append(i.value)
        return columns_list

    def get_rows_number(self, case_id):
        '''
        获取行号
        '''
        num = 1
        cols_data = self.get_columns_value()
        for col_data in cols_data:
            if case_id == col_data:
                return num
            num = num + 1
        return num

    def get_excel_data(self):
        '''
        获取excel里面所有的数据
        '''
        data_list = []
        for i in range(self.get_rows()):
            data_list.append(self.get_rows_value(i + 2))
        return data_list



    # def __get_lines(self):
    #     data =[]
    #     for i in range(self.table.nrows):
    #         data.append(self.table.row_values(i))
    #     return data
    #
    # def __get_columns(self):
    #     data =[]
    #     for i in range(self.table.ncols):
    #         data.append(self.table.col_values(i))
    #     return data
    #
    # def get_data(self):
    #     list =[]
    #     data = self.__get_lines()
    #     for i in range(len(data)):
    #         item = []
    #         for x in range(len(data[i])):
    #             if i>0 and x>0:
    #                 info = data[i][x]
    #                 if info!=None:
    #                     item.append(info)
    #         if item!=None and item!=[]:
    #             list.append(item)
    #     return list


handle_exec = HandleExec()
if __name__ == '__main__':
    ATC.file_path= "xxxxasdasdasd"
    print(ATC.file_path)
    # lo_excl = handle_exec.load_excel("config/data.xlsx")
    # print(handle_exec.get_excel_data())